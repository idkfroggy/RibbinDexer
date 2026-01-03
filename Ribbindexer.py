#!/usr/bin/env python3

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import sqlite3
import csv
import os
import shutil
import json
from datetime import datetime
from pathlib import Path
import threading
import re
import sys


class DocRetrieverApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Ribbindexer")
        
        # Set icon
        try:
            if getattr(sys, 'frozen', False):
                # Running as compiled executable
                base_path = sys._MEIPASS
            else:
                # Running as script
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            icon_path = os.path.join(base_path, 'ribbindexer_logo.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception as e:
            # If icon fails, continue without it even though that would suck because it looks way better with the icon
            print(f"Could not load icon: {e}")
        
        self.root.geometry("950x750")
        
        # Configuration
        self.config_file = 'app_config.json'
        self.db_path = 'document_index.db'
        self.config = self.load_config()
        
        # Database
        self.conn = None
        self.cursor = None
        self.setup_database()
        
        # Variables
        self.drive_location = tk.StringVar(value=self.config.get('drive_location', ''))
        self.output_folder = tk.StringVar(value=self.config.get('output_folder', 'output'))
        self.exclude_folders = tk.StringVar(value=self.config.get('exclude_folders', 'temp,archive,old,backup'))
        self.csv_file_path = tk.StringVar()
        self.search_account = tk.StringVar()
        self.search_firstname = tk.StringVar()
        self.search_lastname = tk.StringVar()
        self.date_from = tk.StringVar()
        self.date_to = tk.StringVar()
        self.bulk_rename_prefix = tk.BooleanVar(value=True)
        self.export_to_excel = tk.BooleanVar(value=True)
        
        # Search history
        self.search_history = self.config.get('search_history', [])
        
        # File type variables for indexing
        self.file_types = {
            'pdf': tk.BooleanVar(value=self.config.get('file_types', {}).get('pdf', True)),
            'xlsx': tk.BooleanVar(value=self.config.get('file_types', {}).get('xlsx', True)),
            'xls': tk.BooleanVar(value=self.config.get('file_types', {}).get('xls', True)),
            'docx': tk.BooleanVar(value=self.config.get('file_types', {}).get('docx', True)),
            'doc': tk.BooleanVar(value=self.config.get('file_types', {}).get('doc', True)),
            'csv': tk.BooleanVar(value=self.config.get('file_types', {}).get('csv', False)),
            'txt': tk.BooleanVar(value=self.config.get('file_types', {}).get('txt', False)),
        }
        
        # File type filters for retrieval
        self.retrieval_filters = {
            'pdf': tk.BooleanVar(value=True),
            'xlsx': tk.BooleanVar(value=True),
            'xls': tk.BooleanVar(value=True),
            'docx': tk.BooleanVar(value=True),
            'doc': tk.BooleanVar(value=True),
            'csv': tk.BooleanVar(value=True),
            'txt': tk.BooleanVar(value=True),
        }
        
        # Content indexing option
        self.index_content = tk.BooleanVar(value=self.config.get('index_content', False))
        
        self.create_gui()
    
    def load_config(self):
        """Load configuration from file"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_config(self):
        """Save configuration to file"""
        config = {
            'drive_location': self.drive_location.get(),
            'output_folder': self.output_folder.get(),
            'exclude_folders': self.exclude_folders.get(),
            'file_types': {k: v.get() for k, v in self.file_types.items()},
            'index_content': self.index_content.get(),
            'search_history': self.search_history[-10:]  # Keep last 10
        }
        with open(self.config_file, 'w') as f:
            json.dump(config, f, indent=2)
    
    def setup_database(self):
        """Initialize database with content search support"""
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        
        # Main files table
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                filepath TEXT,
                file_extension TEXT,
                file_size INTEGER,
                content_text TEXT,
                indexed_at TEXT
            )
        ''')
        
        # Create indexes for fast searching
        self.cursor.execute('CREATE INDEX IF NOT EXISTS idx_filename ON files(filename)')
        self.cursor.execute('CREATE INDEX IF NOT EXISTS idx_extension ON files(file_extension)')
        self.cursor.execute('CREATE INDEX IF NOT EXISTS idx_content ON files(content_text)')
        
        # Create full-text search virtual table for content
        try:
            self.cursor.execute('''
                CREATE VIRTUAL TABLE IF NOT EXISTS files_fts USING fts5(
                    filename, filepath, content_text, content='files', content_rowid='id'
                )
            ''')
        except:
            pass  # FTS5 might not be available
        
        self.conn.commit()
    
    def create_gui(self):
        """For Tabbed Interface"""
        # Create notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create tabs
        self.setup_tab = ttk.Frame(self.notebook)
        self.retrieval_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.setup_tab, text="Setup & Indexing")
        self.notebook.add(self.retrieval_tab, text="Document Retrieval")
        
        self.create_setup_tab()
        self.create_retrieval_tab()
    
    def create_setup_tab(self):
        """For setup and indexing interface"""
        # Main container
        container = ttk.Frame(self.setup_tab, padding=10)
        container.pack(fill="both", expand=True)
        
        # DRIVE CONFIGURATION
        config_frame = ttk.LabelFrame(container, text="Drive Configuration", padding=10)
        config_frame.pack(fill="x", pady=5)
        
        # Drive location
        ttk.Label(config_frame, text="Network Drive Location:").grid(row=0, column=0, sticky="w", pady=5)
        
        drive_frame = ttk.Frame(config_frame)
        drive_frame.grid(row=0, column=1, sticky="ew", pady=5, padx=5)
        
        drive_entry = ttk.Entry(drive_frame, textvariable=self.drive_location, width=50)
        drive_entry.pack(side="left", fill="x", expand=True)
        
        ttk.Button(drive_frame, text="Browse", command=self.browse_drive).pack(side="left", padx=5)
        
        # Example text
        ttk.Label(config_frame, text="Example: Z:\\contracts or \\\\server\\share\\documents", 
                  foreground="gray").grid(row=1, column=1, sticky="w", padx=5)
        
        # Exclude folders
        ttk.Label(config_frame, text="Exclude Folders:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(config_frame, textvariable=self.exclude_folders, width=50).grid(
            row=2, column=1, sticky="w", pady=5, padx=5
        )
        ttk.Label(config_frame, text="Comma-separated folder names to skip (e.g., temp,archive,old)", 
                  foreground="gray").grid(row=3, column=1, sticky="w", padx=5)
        
        # SELECTING FILE TYPES TO INDEX
        types_frame = ttk.LabelFrame(container, text="File Types to Index", padding=10)
        types_frame.pack(fill="x", pady=5)
        
        types_grid = ttk.Frame(types_frame)
        types_grid.pack(fill="x")
        
        col = 0
        for file_type, var in self.file_types.items():
            ttk.Checkbutton(types_grid, text=file_type.upper(), variable=var).grid(
                row=0, column=col, padx=10, pady=5, sticky="w"
            )
            col += 1
        
        # OPTIONS FOR INDEXING
        options_frame = ttk.LabelFrame(container, text="Indexing Options", padding=10)
        options_frame.pack(fill="x", pady=5)
        
        ttk.Checkbutton(
            options_frame, 
            text="Index document contents (slower, enables text search)",
            variable=self.index_content
        ).pack(anchor="w", pady=5)
        
        ttk.Label(
            options_frame,
            text="Note: Content indexing allows searching within documents but takes longer.",
            foreground="gray",
            wraplength=700
        ).pack(anchor="w", pady=5)
        
        # SAVE SETTINGS BUTTON
        ttk.Button(
            container,
            text="💾 Save Settings",
            command=self.save_settings,
            style="Accent.TButton"
        ).pack(pady=10)
        
        # INDEXING SECTION
        index_frame = ttk.LabelFrame(container, text="Start Indexing", padding=10)
        index_frame.pack(fill="both", expand=True, pady=5)
        
        ttk.Label(
            index_frame,
            text="Click 'Start Indexing' to scan the network drive and build the searchable database.",
            wraplength=700
        ).pack(pady=5)
        
        button_frame = ttk.Frame(index_frame)
        button_frame.pack(pady=10)
        
        self.index_btn = ttk.Button(
            button_frame,
            text="▶ Start Full Indexing",
            command=self.start_indexing,
            style="Accent.TButton"
        )
        self.index_btn.pack(side="left", padx=5)
        
        self.update_index_btn = ttk.Button(
            button_frame,
            text="🔄 Update Index (Incremental)",
            command=self.start_incremental_indexing
        )
        self.update_index_btn.pack(side="left", padx=5)
        
        # Progress
        self.index_progress = ttk.Progressbar(index_frame, mode='determinate', length=700)
        self.index_progress.pack(pady=5)
        
        self.index_status = ttk.Label(index_frame, text="Ready to index", foreground="blue")
        self.index_status.pack(pady=5)
        
        # Database info
        info_frame = ttk.LabelFrame(container, text="Database Statistics", padding=10)
        info_frame.pack(fill="x", pady=5)
        
        self.db_info_label = ttk.Label(info_frame, text="Loading...")
        self.db_info_label.pack(pady=5)
        
        ttk.Button(info_frame, text="🔄 Refresh Stats", command=self.update_db_info).pack(pady=5)
        
        # Update stats on load
        self.update_db_info()
    
    def create_retrieval_tab(self):
        """Create document retrieval interface"""
        container = ttk.Frame(self.retrieval_tab, padding=10)
        container.pack(fill="both", expand=True)
        
        # SEARCH OPTIONS
        search_frame = ttk.LabelFrame(container, text="Search Options", padding=10)
        search_frame.pack(fill="x", pady=5)
        
        # CSV Upload
        ttk.Label(search_frame, text="CSV File (Account IDs):").grid(row=0, column=0, sticky="w", pady=5)
        
        csv_frame = ttk.Frame(search_frame)
        csv_frame.grid(row=0, column=1, sticky="ew", pady=5, padx=5)
        
        csv_entry = ttk.Entry(csv_frame, textvariable=self.csv_file_path, width=40)
        csv_entry.pack(side="left", fill="x", expand=True)
        
        ttk.Button(csv_frame, text="Browse", command=self.browse_csv).pack(side="left", padx=5)
        
        # Search history
        ttk.Label(search_frame, text="Recent Searches:").grid(row=0, column=2, sticky="w", pady=5, padx=(20,0))
        self.history_combo = ttk.Combobox(search_frame, values=self.search_history, width=20, state="readonly")
        self.history_combo.grid(row=0, column=3, sticky="w", pady=5, padx=5)
        self.history_combo.bind("<<ComboboxSelected>>", self.load_from_history)
        
        # Account ID search
        ttk.Label(search_frame, text="OR Account ID:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(search_frame, textvariable=self.search_account, width=40).grid(
            row=1, column=1, sticky="w", pady=5, padx=5
        )
        
        # Name search
        ttk.Label(search_frame, text="OR First Name:").grid(row=2, column=0, sticky="w", pady=5)
        ttk.Entry(search_frame, textvariable=self.search_firstname, width=40).grid(
            row=2, column=1, sticky="w", pady=5, padx=5
        )
        
        ttk.Label(search_frame, text="Last Name:").grid(row=3, column=0, sticky="w", pady=5)
        ttk.Entry(search_frame, textvariable=self.search_lastname, width=40).grid(
            row=3, column=1, sticky="w", pady=5, padx=5
        )
        
        # Date filters
        ttk.Label(search_frame, text="Date From (YYYY-MM-DD):").grid(row=4, column=0, sticky="w", pady=5)
        ttk.Entry(search_frame, textvariable=self.date_from, width=20).grid(
            row=4, column=1, sticky="w", pady=5, padx=5
        )
        
        ttk.Label(search_frame, text="Date To (YYYY-MM-DD):").grid(row=4, column=2, sticky="w", pady=5, padx=(20,0))
        ttk.Entry(search_frame, textvariable=self.date_to, width=20).grid(
            row=4, column=3, sticky="w", pady=5, padx=5
        )
        
        # Output folder
        ttk.Label(search_frame, text="Output Folder:").grid(row=5, column=0, sticky="w", pady=5)
        
        output_frame = ttk.Frame(search_frame)
        output_frame.grid(row=5, column=1, sticky="ew", pady=5, padx=5)
        
        ttk.Entry(output_frame, textvariable=self.output_folder, width=40).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(output_frame, text="Browse", command=self.browse_output).pack(side="left", padx=5)
        
        # FILTERS
        filters_frame = ttk.LabelFrame(container, text="Retrieval Filters", padding=10)
        filters_frame.pack(fill="x", pady=5)
        
        # File type filters
        ttk.Label(filters_frame, text="File Types to Retrieve:").pack(anchor="w", pady=5)
        filter_grid = ttk.Frame(filters_frame)
        filter_grid.pack(fill="x")
        
        col = 0
        for file_type, var in self.retrieval_filters.items():
            ttk.Checkbutton(filter_grid, text=file_type.upper(), variable=var).grid(
                row=0, column=col, padx=10, pady=5, sticky="w"
            )
            col += 1
        
        # RETRIEVAL OPTIONS
        options_frame = ttk.LabelFrame(container, text="Options", padding=10)
        options_frame.pack(fill="x", pady=5)
        
        ttk.Checkbutton(
            options_frame,
            text="Prefix retrieved files with account/search term",
            variable=self.bulk_rename_prefix
        ).pack(anchor="w", pady=2)
        
        ttk.Checkbutton(
            options_frame,
            text="Export manifest to Excel (.xlsx) format",
            variable=self.export_to_excel
        ).pack(anchor="w", pady=2)
        
        # RETRIEVE BUTTON AND DENNIS BUTTON
        button_container = ttk.Frame(container)
        button_container.pack(pady=10)
        
        self.retrieve_btn = ttk.Button(
            button_container,
            text="Search & Retrieve Documents",
            command=self.start_retrieval,
            style="Accent.TButton"
        )
        self.retrieve_btn.pack(side="left", padx=5)
        
        ttk.Button(
            button_container,
            text="❓ DENNIS BUTTON",
            command=self.show_help
        ).pack(side="left", padx=5)
        
        # Progress
        self.retrieve_progress = ttk.Progressbar(container, mode='determinate', length=700)
        self.retrieve_progress.pack(pady=5)
        
        self.retrieve_status = ttk.Label(container, text="Ready to retrieve", foreground="blue")
        self.retrieve_status.pack(pady=5)
        
        # RESULTS
        results_frame = ttk.LabelFrame(container, text="Search Results & Preview", padding=10)
        results_frame.pack(fill="both", expand=True, pady=5)
        
        # Split results and preview
        paned = ttk.PanedWindow(results_frame, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True)
        
        # Results list
        results_left = ttk.Frame(paned)
        paned.add(results_left, weight=1)
        
        ttk.Label(results_left, text="Results Summary:").pack(anchor="w")
        self.results_text = scrolledtext.ScrolledText(results_left, height=15, width=50)
        self.results_text.pack(fill="both", expand=True)
        
        # Preview pane
        preview_right = ttk.Frame(paned)
        paned.add(preview_right, weight=1)
        
        ttk.Label(preview_right, text="File Preview:").pack(anchor="w")
        self.preview_text = scrolledtext.ScrolledText(preview_right, height=15, width=50, wrap=tk.WORD)
        self.preview_text.pack(fill="both", expand=True)
        self.preview_text.insert(tk.END, "Preview will appear here when files are found...")
    
    def browse_drive(self):
        """Browse drive location"""
        folder = filedialog.askdirectory(title="Select Network Drive or Folder")
        if folder:
            self.drive_location.set(folder)
    
    def browse_csv(self):
        """Browse CSV file"""
        filename = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_file_path.set(filename)
    
    def browse_output(self):
        """Browse output folder"""
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder.set(folder)
    
    def save_settings(self):
        """Save current settings"""
        self.save_config()
        messagebox.showinfo("Settings Saved", "Configuration has been saved successfully!")
    
    def load_from_history(self, event=None):
        """Load search from history"""
        selected = self.history_combo.get()
        if selected:
            self.search_account.set(selected)
    
    def add_to_history(self, term):
        """Add search term to history"""
        if term and term not in self.search_history:
            self.search_history.insert(0, term)
            self.search_history = self.search_history[:10]  # Keep last 10
            self.history_combo['values'] = self.search_history
            self.save_config()
    
    def show_help(self):
        """Show comprehensive how-to guide for using the application"""
        help_window = tk.Toplevel(self.root)
        help_window.title("How to Use Ribbindexer - DENNIS BUTTON Help")
        help_window.geometry("800x600")
        
        # Create scrolled text widget for help content
        help_text = scrolledtext.ScrolledText(
            help_window,
            wrap=tk.WORD,
            width=90,
            height=35,
            font=("Segoe UI", 10)
        )
        help_text.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Help content
        help_content = """
                    SPECIAL THANKS TO MIKE!

OVERVIEW
-----------
Ribbindexer helps you index and retrieve documents from network drives.
It has two main functions:
  1. Index documents from a network drive (Setup & Indexing tab)
  2. Search and retrieve documents by account ID or name (Document Retrieval tab)


STEP 1: SETUP & INDEXING TAB
--------------------------------

1. Configure Drive Location:
   • Click "Browse" next to "Network Drive Location"
   • Select your network drive or folder (e.g., Z:\\contracts)
   • Or type the path directly (e.g., \\\\server\\share\\documents)

2. Set Exclude Folders (Optional):
   • Enter folder names to skip during indexing
   • Separate with commas (e.g., temp,archive,old,backup)
   • This speeds up indexing by skipping unnecessary folders

3. Select File Types to Index:
   • Check the boxes for file types you want to search
   • Common types: PDF, XLSX, XLS, DOCX, DOC
   • Optional: CSV, TXT

4. Indexing Options:
   • Enable "Index document contents" for text search inside files
   • Warning: This is slower but allows searching document text

5. Save Settings:
   • Click "Save Settings" to save your configuration

6. Start Indexing:
   • Click "Start Full Indexing" for first-time indexing
   • Click "Update Index (Incremental)" to update existing index
   • Wait for indexing to complete (progress bar shows status)

7. Check Database Statistics:
   • View total files indexed by type
   • Click "Refresh Stats" to update


STEP 2: DOCUMENT RETRIEVAL TAB
----------------------------------

You can search in THREE ways:

OPTION A - Upload CSV File:
   • Click "Browse" next to "CSV File (Account IDs)"
   • Select a CSV file containing account IDs
   • File should have account numbers (one per row or in a column)
   • System will search for all accounts in the file

OPTION B - Single Account ID:
   • Type an account number in "OR Account ID" field
   • Example: 123456

OPTION C - Search by Name:
   • Enter "First Name" and/or "Last Name"
   • System searches for documents containing these names

8. Optional Filters:
   • Date From/To: Filter by date in YYYY-MM-DD format
     Example: 2024-01-01 to 2024-12-31
   • File Types to Retrieve: Select which file types to copy

9. Set Output Folder:
   • Click "Browse" next to "Output Folder"
   • Choose where retrieved documents will be saved
   • Default: "output" folder in program directory

10. Options:
    • "Prefix retrieved files with account/search term"
      → Adds account ID to filename for easy identification
    • "Export manifest to Excel (.xlsx) format"
      → Creates Excel spreadsheet listing all retrieved files

11. Search & Retrieve:
    • Click "Search & Retrieve Documents" button
    • Watch progress bar for status
    • Results appear in the "Search Results & Preview" section


OUTPUT STRUCTURE
-------------------
Retrieved documents are saved in a timestamped folder:
  output/retrieval_YYYYMMDD_HHMMSS/

Contains:
  • All matching documents (copied from network drive)
  • MANIFEST.txt - Text list of retrieved files
  • MANIFEST.xlsx - Excel spreadsheet (if option enabled)


TIPS & BEST PRACTICES
------------------------

✓ Index First: Always index your network drive before searching
✓ Regular Updates: Run incremental indexing weekly or when files change
✓ CSV Format: Account IDs in CSV can be in any column, one per row
✓ Date Format: Use YYYY-MM-DD format (e.g., 2024-03-15)
✓ File Naming: Documents must contain account ID or name in filename
✓ Recent Searches: Use the "Recent Searches" dropdown for quick access
✓ Prefixing: Enable prefixing to avoid duplicate filenames
✓ Preview: Check the preview pane to verify found documents


TROUBLESHOOTING
------------------

No Results Found?
  • Verify indexing completed successfully
  • Check that account ID/name appears in document filenames
  • Try broadening your search (remove date filters)
  • Verify file types are selected in retrieval filters

Indexing Failed?
  • Check network drive path is accessible
  • Check exclude folders aren't blocking needed files

Slow Performance?
  • Disable "Index document contents" if not needed
  • Add temporary folders to exclude list
  • Use incremental indexing instead of full indexing

"""
        
        # Insert help content
        help_text.insert(tk.END, help_content)
        help_text.config(state="disabled")  # Make read-only
        
        # Add close button
        ttk.Button(
            help_window,
            text="Close",
            command=help_window.destroy
        ).pack(pady=10)

    
    def get_selected_extensions(self):
        """Get list of selected file extensions"""
        extensions = []
        for ext, var in self.file_types.items():
            if var.get():
                extensions.append(f'.{ext}')
        return extensions
    
    def extract_text_from_file(self, filepath):
        try:
            ext = os.path.splitext(filepath)[1].lower()
            
            if ext == '.txt':
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            
            elif ext == '.csv':
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            
            elif ext == '.pdf':
                try:
                    import PyPDF2
                    text = []
                    with open(filepath, 'rb') as f:
                        pdf_reader = PyPDF2.PdfReader(f)
                        for page in pdf_reader.pages:
                            text.append(page.extract_text())
                    return '\n'.join(text)
                except ImportError:
                    return "[PDF - Install PyPDF2: pip install PyPDF2]"
                except Exception as e:
                    return f"[PDF extraction error: {str(e)}]"
            
            elif ext == '.docx':
                try:
                    from docx import Document
                    doc = Document(filepath)
                    text = []
                    for paragraph in doc.paragraphs:
                        text.append(paragraph.text)
                    return '\n'.join(text)
                except ImportError:
                    return "[DOCX - Install python-docx: pip install python-docx]"
                except Exception as e:
                    return f"[DOCX extraction error: {str(e)}]"
            
            elif ext == '.xlsx':
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filepath, read_only=True, data_only=True)
                    text = []
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            row_text = ' '.join(str(cell) for cell in row if cell is not None)
                            if row_text.strip():
                                text.append(row_text)
                    wb.close()
                    return '\n'.join(text)
                except ImportError:
                    return "[XLSX - Install openpyxl: pip install openpyxl]"
                except Exception as e:
                    return f"[XLSX extraction error: {str(e)}]"
            
            elif ext == '.xls':
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(filepath)
                    text = []
                    for sheet in workbook.sheets():
                        for row_idx in range(sheet.nrows):
                            row = sheet.row_values(row_idx)
                            row_text = ' '.join(str(cell) for cell in row if cell)
                            if row_text.strip():
                                text.append(row_text)
                    return '\n'.join(text)
                except ImportError:
                    return "[XLS - Install xlrd: pip install xlrd]"
                except Exception as e:
                    return f"[XLS extraction error: {str(e)}]"
            
            else:
                return ""
        
        except Exception as e:
            return f"[Error extracting text: {str(e)}]"
    
    def start_indexing(self):
        """Start indexing process"""
        drive = self.drive_location.get().strip()
        
        if not drive:
            messagebox.showerror("Error", "Please enter a drive location")
            return
        
        if not os.path.exists(drive):
            messagebox.showerror("Error", "Drive location does not exist")
            return
        
        extensions = self.get_selected_extensions()
        if not extensions:
            messagebox.showerror("Error", "Please select at least one file type")
            return
        
        # Disable button
        self.index_btn.config(state="disabled")
        self.index_status.config(text="Indexing in progress...", foreground="orange")
        
        # Run in background thread
        thread = threading.Thread(
            target=self.index_directory,
            args=(drive, extensions, self.index_content.get(), False)
        )
        thread.daemon = True
        thread.start()
    
    def start_incremental_indexing(self):
        """Start incremental indexing process"""
        drive = self.drive_location.get().strip()
        
        if not drive:
            messagebox.showerror("Error", "Please enter a drive location")
            return
        
        if not os.path.exists(drive):
            messagebox.showerror("Error", "Drive location does not exist")
            return
        
        extensions = self.get_selected_extensions()
        if not extensions:
            messagebox.showerror("Error", "Please select at least one file type")
            return
        
        # Disable buttons
        self.index_btn.config(state="disabled")
        self.update_index_btn.config(state="disabled")
        self.index_status.config(text="Incremental indexing in progress...", foreground="orange")
        
        # Run in background thread
        thread = threading.Thread(
            target=self.index_directory,
            args=(drive, extensions, self.index_content.get(), True)
        )
        thread.daemon = True
        thread.start()
    
    def index_directory(self, directory, extensions, index_content, incremental=False):
        """Index directory in chunks"""
        try:
            # Get exclude folders
            exclude_folders = [f.strip().lower() for f in self.exclude_folders.get().split(',') if f.strip()]
            
            # Clear existing entries unless incremental
            if not incremental:
                self.cursor.execute("DELETE FROM files")
                self.conn.commit()
            else:
                # Get existing indexed files for comparison
                self.cursor.execute("SELECT filepath FROM files")
                existing_files = {row[0] for row in self.cursor.fetchall()}
            
            # Collect all files
            self.index_status.config(text="Scanning directory structure...")
            all_files = []
            
            for root, dirs, files in os.walk(directory):
                # Skip excluded folders
                dirs[:] = [d for d in dirs if d.lower() not in exclude_folders]
                
                for filename in files:
                    ext = os.path.splitext(filename)[1].lower()
                    if ext in extensions:
                        filepath = os.path.join(root, filename)
                        
                        # For incremental, only add new files
                        if incremental:
                            if filepath not in existing_files:
                                all_files.append(filepath)
                        else:
                            all_files.append(filepath)
            
            total_files = len(all_files)
            if total_files == 0:
                self.index_status.config(text="No files found", foreground="red")
                self.index_btn.config(state="normal")
                return
            
            # Process in chunks because I don't want to crash everything
            chunk_size = 50  
            indexed_count = 0
            
            for i in range(0, total_files, chunk_size):
                chunk = all_files[i:i + chunk_size]
                
                for filepath in chunk:
                    try:
                        filename = os.path.basename(filepath)
                        extension = os.path.splitext(filename)[1].lower()
                        size = os.path.getsize(filepath)
                        
                        # Extract content if enabled (WARN THEM THAT THIS WILL PROLONG THE PROCESS)
                        content_text = ""
                        if index_content:
                            content_text = self.extract_text_from_file(filepath)
                        
                        self.cursor.execute(
                            '''INSERT INTO files (filename, filepath, file_extension, file_size, content_text, indexed_at) 
                               VALUES (?, ?, ?, ?, ?, ?)''',
                            (filename, filepath, extension, size, content_text, datetime.now().isoformat())
                        )
                        indexed_count += 1
                    
                    except Exception as e:
                        print(f"Error indexing {filepath}: {e}")
                
                # Commit chunk
                self.conn.commit()
                
                # Progress update
                progress = (indexed_count / total_files) * 100
                self.index_progress['value'] = progress
                
                status_msg = f"Indexed {indexed_count:,} of {total_files:,} files ({progress:.1f}%)"
                if index_content:
                    status_msg += " - Extracting content..."
                
                self.index_status.config(text=status_msg)
                self.root.update_idletasks()
            
            # Complete
            self.index_progress['value'] = 100
            self.index_status.config(
                text=f"✓ Complete! Indexed {indexed_count:,} files",
                foreground="green"
            )
            self.update_db_info()
        
        except Exception as e:
            self.index_status.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Indexing Error", str(e))
        
        finally:
            self.index_btn.config(state="normal")
            self.update_index_btn.config(state="normal")
            self.index_progress['value'] = 0
    
    def start_retrieval(self):
        """Start document retrieval"""
        # Determine search mode
        csv_file = self.csv_file_path.get().strip()
        account = self.search_account.get().strip()
        firstname = self.search_firstname.get().strip()
        lastname = self.search_lastname.get().strip()
        
        if not any([csv_file, account, firstname, lastname]):
            messagebox.showerror("Error", "Please provide search criteria (CSV, Account ID, or Name)")
            return
        
        # Disable button
        self.retrieve_btn.config(state="disabled")
        self.retrieve_status.config(text="Searching...", foreground="orange")
        self.results_text.delete(1.0, tk.END)
        
        # Run in background
        thread = threading.Thread(target=self.retrieve_documents)
        thread.daemon = True
        thread.start()
    
    def retrieve_documents(self):
        """Retrieve documents based on search criteria"""
        try:
            import hashlib
            
            search_terms = []
            
            # Get search terms from various sources
            csv_file = self.csv_file_path.get().strip()
            if csv_file and os.path.exists(csv_file):
                with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    next(reader, None)  # Skip header
                    for row in reader:
                        if row:
                            term = row[0].strip()
                            search_terms.append(('account', term))
                            self.add_to_history(term)
            
            account = self.search_account.get().strip()
            if account:
                search_terms.append(('account', account))
                self.add_to_history(account)
            
            firstname = self.search_firstname.get().strip()
            lastname = self.search_lastname.get().strip()
            if firstname or lastname:
                name_query = f"{firstname} {lastname}".strip()
                search_terms.append(('name', name_query))
                self.add_to_history(name_query)
            
            if not search_terms:
                self.retrieve_status.config(text="No search terms provided", foreground="red")
                return
            
            # Get date filters
            date_from = self.date_from.get().strip()
            date_to = self.date_to.get().strip()
            
            # Get file type filters
            allowed_extensions = [f'.{ext}' for ext, var in self.retrieval_filters.items() if var.get()]
            
            # Create output folder
            output_dir = self.output_folder.get()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(output_dir, f'retrieval_{timestamp}')
            os.makedirs(output_path, exist_ok=True)
            
            # Track all found files for CSV export
            found_files_data = []
            
            # Search and retrieve
            manifest_data = []
            manifest_data.append(["DOCUMENT RETRIEVAL MANIFEST"])
            manifest_data.append([f"Generated: {datetime.now()}"])
            manifest_data.append([f"Total Search Terms: {len(search_terms)}"])
            manifest_data.append(["=" * 70])
            manifest_data.append([])
            
            results_summary = []
            missing_files = []
            duplicates_detected = []
            total_found = 0
            processed = 0
            file_hashes = {}  # For duplicate detection
            
            # Clear preview
            self.preview_text.delete(1.0, tk.END)
            
            for search_type, term in search_terms:
                # Build query with date filter if provided
                if search_type == 'account':
                    query = "SELECT * FROM files WHERE (filename LIKE ? OR content_text LIKE ?)"
                    params = [f'%{term}%', f'%{term}%']
                else:  # name search
                    query = "SELECT * FROM files WHERE content_text LIKE ?"
                    params = [f'%{term}%']
                
                # Add date filter if provided
                if date_from or date_to:
                    query += " AND indexed_at >= ? AND indexed_at <= ?"
                    params.append(date_from if date_from else '1900-01-01')
                    params.append(date_to if date_to else '2100-12-31')
                
                self.cursor.execute(query, params)
                results = self.cursor.fetchall()
                
                # Filter by file type
                filtered_results = [r for r in results if r['file_extension'] in allowed_extensions]
                
                manifest_data.append([f"Search Term: {term} ({search_type})"])
                results_summary.append(f"{term}: {len(filtered_results)} file(s)")
                
                if not filtered_results:
                    manifest_data.append(["  ✗ NOT FOUND"])
                    missing_files.append({'search_term': term, 'search_type': search_type})
                else:
                    total_found += len(filtered_results)
                    for file in filtered_results:
                        try:
                            # Calculate file hash for duplicate detection
                            file_hash = None
                            try:
                                with open(file['filepath'], 'rb') as f:
                                    file_hash = hashlib.md5(f.read()).hexdigest()
                            except:
                                pass
                            
                            # Check for duplicate
                            is_duplicate = False
                            if file_hash and file_hash in file_hashes:
                                is_duplicate = True
                                duplicates_detected.append({
                                    'original': file_hashes[file_hash],
                                    'duplicate': file['filename'],
                                    'search_term': term
                                })
                            else:
                                if file_hash:
                                    file_hashes[file_hash] = file['filename']
                            
                            # Prepare filename (with optional prefix)
                            filename = file['filename']
                            if self.bulk_rename_prefix.get():
                                base, ext = os.path.splitext(filename)
                                filename = f"{term}_{base}{ext}"
                            
                            dest = os.path.join(output_path, filename)
                            
                            # Handle name conflicts
                            if os.path.exists(dest):
                                base, ext = os.path.splitext(filename)
                                counter = 1
                                while os.path.exists(dest):
                                    dest = os.path.join(output_path, f"{base}_{counter}{ext}")
                                    counter += 1
                            
                            # Copy file
                            shutil.copy2(file['filepath'], dest)
                            
                            # Record in manifest
                            dup_marker = " [DUPLICATE]" if is_duplicate else ""
                            manifest_data.append([f"  ✓ {filename}{dup_marker}"])
                            manifest_data.append([f"    Source: {file['filepath']}"])
                            manifest_data.append([f"    Size: {file['file_size']:,} bytes"])
                            
                            # Add to found files data for CSV
                            found_files_data.append({
                                'search_term': term,
                                'search_type': search_type,
                                'filename': file['filename'],
                                'filepath': file['filepath'],
                                'file_extension': file['file_extension'],
                                'file_size': file['file_size'],
                                'output_filename': os.path.basename(dest),
                                'is_duplicate': is_duplicate
                            })
                        
                        except Exception as e:
                            manifest_data.append([f"  ✗ ERROR: {str(e)}"])
                
                manifest_data.append([])
                
                # Update progress
                processed += 1
                progress = (processed / len(search_terms)) * 100
                self.retrieve_progress['value'] = progress
                self.retrieve_status.config(
                    text=f"Processing {processed} of {len(search_terms)} searches..."
                )
                self.root.update_idletasks()
            
            # Save manifest
            manifest_data.append([f"SUMMARY: Retrieved {total_found} files for {len(search_terms)} search terms"])
            if duplicates_detected:
                manifest_data.append([f"Duplicates detected: {len(duplicates_detected)}"])
            
            # Export manifest as text
            manifest_txt_path = os.path.join(output_path, 'MANIFEST.txt')
            with open(manifest_txt_path, 'w', encoding='utf-8') as f:
                for row in manifest_data:
                    f.write(' '.join(str(item) for item in row) + '\n')
            
            # Export manifest to Excel if enabled
            if self.export_to_excel.get():
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Retrieval Manifest"
                    
                    for row_data in manifest_data:
                        ws.append(row_data)
                    
                    manifest_xlsx_path = os.path.join(output_path, 'MANIFEST.xlsx')
                    wb.save(manifest_xlsx_path)
                except ImportError:
                    pass  # Skip if openpyxl not available
                except Exception as e:
                    print(f"Error creating Excel manifest: {e}")
            
            # Save missing files to CSV if any
            if missing_files:
                missing_csv_path = os.path.join(output_path, 'missing_media.csv')
                with open(missing_csv_path, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['search_term', 'search_type'])
                    writer.writeheader()
                    writer.writerows(missing_files)
            
            # Save found files to CSV
            if found_files_data:
                found_csv_path = os.path.join(output_path, 'found_files.csv')
                with open(found_csv_path, 'w', encoding='utf-8', newline='') as f:
                    fieldnames = ['search_term', 'search_type', 'filename', 'filepath', 
                                  'file_extension', 'file_size', 'output_filename', 'is_duplicate']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(found_files_data)
            
            # Save duplicates report if any
            if duplicates_detected:
                dup_csv_path = os.path.join(output_path, 'duplicates_detected.csv')
                with open(dup_csv_path, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=['original', 'duplicate', 'search_term'])
                    writer.writeheader()
                    writer.writerows(duplicates_detected)
            
            # Display results
            self.results_text.delete(1.0, tk.END)
            self.results_text.insert(tk.END, "Search Results:\n\n")
            for result in results_summary:
                self.results_text.insert(tk.END, f"  {result}\n")
            self.results_text.insert(tk.END, f"\nTotal Files Retrieved: {total_found}\n")
            if missing_files:
                self.results_text.insert(tk.END, f"Missing Files: {len(missing_files)} (saved to missing_media.csv)\n")
            if duplicates_detected:
                self.results_text.insert(tk.END, f"Duplicates Detected: {len(duplicates_detected)} (saved to duplicates_detected.csv)\n")
            self.results_text.insert(tk.END, f"\nOutput Folder: {output_path}\n")
            self.results_text.insert(tk.END, f"Found files list: found_files.csv\n")
            
            # Show preview of first found file
            if found_files_data:
                try:
                    first_file = found_files_data[0]['filepath']
                    preview_text = self.extract_text_from_file(first_file)
                    if preview_text:
                        preview_text = preview_text[:2000]  # Limit to 2000 chars
                        self.preview_text.delete(1.0, tk.END)
                        self.preview_text.insert(tk.END, f"Preview of: {found_files_data[0]['filename']}\n\n")
                        self.preview_text.insert(tk.END, preview_text)
                        self.preview_text.insert(tk.END, "\n\n... (truncated)")
                except Exception as e:
                    self.preview_text.delete(1.0, tk.END)
                    self.preview_text.insert(tk.END, f"Preview error: {e}")
            
            # Complete
            self.retrieve_progress['value'] = 100
            self.retrieve_status.config(
                text=f"✓ Complete! Retrieved {total_found} files",
                foreground="green"
            )
            
            # Ask to open folder
            if messagebox.askyesno("Success", f"Retrieved {total_found} files!\n\nOpen output folder?"):
                if os.name == 'nt':
                    os.startfile(output_path)
                else:
                    os.system(f'xdg-open "{output_path}"')
        
        except Exception as e:
            self.retrieve_status.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Retrieval Error", str(e))
        
        finally:
            self.retrieve_btn.config(state="normal")
            self.retrieve_progress['value'] = 0
    
    def update_db_info(self):
        """Update database statistics"""
        try:
            self.cursor.execute("SELECT COUNT(*) as count FROM files")
            total_files = self.cursor.fetchone()['count']
            
            self.cursor.execute("SELECT COUNT(DISTINCT file_extension) as count FROM files")
            total_types = self.cursor.fetchone()['count']
            
            self.cursor.execute("SELECT indexed_at FROM files ORDER BY indexed_at DESC LIMIT 1")
            row = self.cursor.fetchone()
            last_indexed = row['indexed_at'] if row else "Never"
            
            self.cursor.execute("SELECT SUM(file_size) as total FROM files")
            total_size = self.cursor.fetchone()['total'] or 0
            size_mb = total_size / (1024 * 1024)
            
            info_text = (
                f"Total Files: {total_files:,} | "
                f"File Types: {total_types} | "
                f"Total Size: {size_mb:.2f} MB | "
                f"Last Indexed: {last_indexed}"
            )
            self.db_info_label.config(text=info_text)
        
        except Exception as e:
            self.db_info_label.config(text=f"Error: {e}")
    
    def on_closing(self):
        """Clean up on exit"""
        self.save_config()
        if self.conn:
            self.conn.close()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = DocRetrieverApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == '__main__':
    main()